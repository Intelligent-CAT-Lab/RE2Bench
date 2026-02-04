import inspect
import json
import os
from datetime import datetime

def custom_serializer(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    return str(obj)


def recursive_object_seralizer(obj, visited):
    seralized_dict = {}
    keys = list(obj.__dict__)
    for k in keys:
        if id(obj.__dict__[k]) in visited:
            seralized_dict[k] = "<RECURSIVE {}>".format(obj.__dict__[k])
            continue
        if isinstance(obj.__dict__[k], (float, int, str, bool, type(None))):
            seralized_dict[k] = obj.__dict__[k]
        elif isinstance(obj.__dict__[k], tuple):
            ## handle tuple
            seralized_dict[k] = obj.__dict__[k]
        elif isinstance(obj.__dict__[k], set):
            ## handle set
            seralized_dict[k] = obj.__dict__[k]
        elif isinstance(obj.__dict__[k], list):
            ## handle list
            seralized_dict[k] = obj.__dict__[k]
        elif hasattr(obj.__dict__[k], '__dict__'):
            ## handle object
            visited.append(id(obj.__dict__[k]))
            seralized_dict[k] = obj.__dict__[k]
        elif isinstance(obj.__dict__[k], dict):
            visited.append(id(obj.__dict__[k]))
            seralized_dict[k] = obj.__dict__[k]
        elif callable(obj.__dict__[k]):
            ## handle function
            if hasattr(obj.__dict__[k], '__name__'):
                seralized_dict[k] = "<function {}>".format(obj.__dict__[k].__name__)
        else:
            seralized_dict[k] = str(obj.__dict__[k])
    return seralized_dict

def inspect_code(func):
   def wrapper(*args, **kwargs):
       visited = []
       json_base = "/home/changshu/ClassEval/data/benchmark_solution_code/input-output/"
       if not os.path.exists(json_base):
           os.mkdir(json_base)
       jsonl_path = json_base + "/ExcelProcessor.jsonl"
       para_dict = {"name": func.__name__}
       args_names = inspect.getfullargspec(func).args
       if len(args) > 0 and hasattr(args[0], '__dict__') and args_names[0] == 'self':
           ## 'self'
           self_args = args[0]
           para_dict['self'] = recursive_object_seralizer(self_args, [id(self_args)])
       else:
           para_dict['self'] = {}
       if len(args) > 0 :
           if args_names[0] == 'self':
               other_args = {}
               for m,n in zip(args_names[1:], args[1:]):
                   other_args[m] = n
           else:
               other_args = {}
               for m,n in zip(args_names, args):
                   other_args[m] = n
           
           para_dict['args'] = other_args
       else:
           para_dict['args'] = {}
       if kwargs:
           para_dict['kwargs'] = kwargs
       else:
           para_dict['kwargs'] = {}
          
       result = func(*args, **kwargs)
       para_dict["return"] = result
       with open(jsonl_path, 'a') as f:
           f.write(json.dumps(para_dict, default=custom_serializer) + "\n")
       return result
   return wrapper


'''
# This is a class for processing excel files, including readring and writing excel data, as well as processing specific operations and saving as a new excel file.

import openpyxl


class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """


    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        >>> processor = ExcelProcessor()
        >>> new_data = [
        >>>     ('Name', 'Age', 'Country'),
        >>>     ('John', 25, 'USA'),
        >>>     ('Alice', 30, 'Canada'),
        >>>     ('Bob', 35, 'Australia'),
        >>>     ('Julia', 28, 'Germany')
        >>> ]
        >>> data = processor.write_excel(new_data, 'test_data.xlsx')
        """


    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """


'''

import openpyxl


class ExcelProcessor:
    def __init__(self):
        pass

    @inspect_code
    def read_excel(self, file_name):
        data = []
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            workbook.close()
            return data
        except:
            return None

    @inspect_code
    def write_excel(self, data, file_name):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            workbook.close()
            return 1
        except:
            return 0

    @inspect_code
    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if data is None or N >= len(data[0]):
            return 0
        new_data = []
        for row in data:
            new_row = list(row[:])
            if not str(row[N]).isdigit():
                new_row.append(str(row[N]).upper())
            else:
                new_row.append(row[N])
            new_data.append(new_row)
        new_file_name = save_file_name.split('.')[0] + '_process.xlsx'
        success = self.write_excel(new_data, new_file_name)
        return success, new_file_name



import unittest
import os


class ExcelProcessorTestReadExcel(unittest.TestCase):
    def test_read_excel_1(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Age', 'Country'],
                ['John', 25, 'USA'],
                ['Alice', 30, 'Canada'],
                ['Bob', 35, 'Australia']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        data = processor.read_excel(self.test_file_name)
        expected_data = [
            ('Name', 'Age', 'Country'),
            ('John', 25, 'USA'),
            ('Alice', 30, 'Canada'),
            ('Bob', 35, 'Australia')
        ]
        self.assertEqual(data, expected_data)

    def test_read_excel_2(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Age'],
                ['John', 25],
                ['Alice', 30],
                ['Bob', 35]]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        data = processor.read_excel(self.test_file_name)
        expected_data = [
            ('Name', 'Age'),
            ('John', 25),
            ('Alice', 30),
            ('Bob', 35)
        ]
        self.assertEqual(data, expected_data)

    def test_read_excel_3(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name'],
                ['John'],
                ['Alice'],
                ['Bob']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        data = processor.read_excel(self.test_file_name)
        expected_data = [
            ('Name',),
            ('John',),
            ('Alice',),
            ('Bob',)
        ]
        self.assertEqual(data, expected_data)

    def test_read_excel_4(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Country'],
                ['John', 'USA'],
                ['Alice', 'Canada'],
                ['Bob', 'Australia']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        data = processor.read_excel(self.test_file_name)
        expected_data = [
            ('Name', 'Country'),
            ('John', 'USA'),
            ('Alice', 'Canada'),
            ('Bob', 'Australia')
        ]
        self.assertEqual(data, expected_data)

    def test_read_excel_5(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Country'],
                ['John', 'USA']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        data = processor.read_excel(self.test_file_name)
        expected_data = [
            ('Name', 'Country'),
            ('John', 'USA')
        ]
        self.assertEqual(data, expected_data)

    def test_read_excel_6(self):
        self.test_file_name = ''
        processor = ExcelProcessor()
        res = processor.read_excel(self.test_file_name)
        self.assertEqual(res, None)


class ExcelProcessorTestWriteExcel(unittest.TestCase):
    def test_write_excel_1(self):
        processor = ExcelProcessor()
        new_data = [
            ('Name', 'Age', 'Country'),
            ('John', 25, 'USA'),
            ('Alice', 30, 'Canada'),
            ('Bob', 35, 'Australia'),
            ('Julia', 28, 'Germany')
        ]
        save_file_name = 'test_output.xlsx'
        success = processor.write_excel(new_data, save_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.exists(save_file_name))
        saved_data = processor.read_excel(save_file_name)
        self.assertEqual(saved_data, new_data)
        os.remove(save_file_name)

    def test_write_excel_2(self):
        processor = ExcelProcessor()
        new_data = [
            ('Name', 'Age'),
            ('John', 25),
            ('Alice', 30),
            ('Bob', 35),
            ('Julia', 28)
        ]
        save_file_name = 'test_output.xlsx'
        success = processor.write_excel(new_data, save_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.exists(save_file_name))
        saved_data = processor.read_excel(save_file_name)
        self.assertEqual(saved_data, new_data)
        os.remove(save_file_name)

    def test_write_excel_3(self):
        processor = ExcelProcessor()
        new_data = [
            ('Name', 'Age', 'Country'),
            ('John', 25, 'USA'),
            ('Alice', 30, 'Canada'),
            ('Bob', 35, 'Australia')
        ]
        save_file_name = 'test_output.xlsx'
        success = processor.write_excel(new_data, save_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.exists(save_file_name))
        saved_data = processor.read_excel(save_file_name)
        self.assertEqual(saved_data, new_data)
        os.remove(save_file_name)

    def test_write_excel_4(self):
        processor = ExcelProcessor()
        new_data = [
            ('Name', 'Age', 'Country'),
            ('John', 25, 'USA'),
            ('Alice', 30, 'Canada')
        ]
        save_file_name = 'test_output.xlsx'
        success = processor.write_excel(new_data, save_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.exists(save_file_name))
        saved_data = processor.read_excel(save_file_name)
        self.assertEqual(saved_data, new_data)
        os.remove(save_file_name)

    def test_write_excel_5(self):
        processor = ExcelProcessor()
        new_data = [
            ('Name', 'Age', 'Country'),
            ('John', 25, 'USA')
        ]
        save_file_name = 'test_output.xlsx'
        success = processor.write_excel(new_data, save_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.exists(save_file_name))
        saved_data = processor.read_excel(save_file_name)
        self.assertEqual(saved_data, new_data)
        os.remove(save_file_name)

    def test_write_excel_6(self):
        processor = ExcelProcessor()
        new_data = [
            ('Name', 'Age', 'Country'),
            ('John', 25, 'USA')
        ]
        save_file_name = ''
        success = processor.write_excel(new_data, save_file_name)
        self.assertEqual(success, 0)


class ExcelProcessorTestProcessExcelData(unittest.TestCase):
    def test_process_excel_data_1(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Age', 'Country'],
                ['John', 25, 'USA'],
                ['Alice', 30, 'Canada'],
                ['Bob', 35, 'Australia']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        N = 1
        success, output_file = processor.process_excel_data(N, self.test_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.isfile(output_file))
        processed_data = processor.read_excel(output_file)
        expected_processed_data = [
            ('Name', 'Age', 'Country', 'AGE'),
            ('John', 25, 'USA', 25),
            ('Alice', 30, 'Canada', 30),
            ('Bob', 35, 'Australia', 35)
        ]
        self.assertEqual(processed_data, expected_processed_data)
        os.remove(output_file)

    def test_process_excel_data_2(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Age', 'Country'],
                ['John', 25, 'USA'],
                ['Alice', 30, 'Canada'],
                ['Bob', 35, 'Australia']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        N = 0
        success, output_file = processor.process_excel_data(N, self.test_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.isfile(output_file))
        processed_data = processor.read_excel(output_file)
        expected_processed_data = [
            ('Name', 'Age', 'Country', 'NAME'),
            ('John', 25, 'USA', 'JOHN'),
            ('Alice', 30, 'Canada', 'ALICE'),
            ('Bob', 35, 'Australia', 'BOB')
        ]
        self.assertEqual(processed_data, expected_processed_data)
        os.remove(output_file)

    def test_process_excel_data_3(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Age', 'Country'],
                ['John', 25, 'USA'],
                ['Alice', 30, 'Canada'],
                ['Bob', 35, 'Australia']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        N = 2
        success, output_file = processor.process_excel_data(N, self.test_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.isfile(output_file))
        processed_data = processor.read_excel(output_file)
        expected_processed_data = [
            ('Name', 'Age', 'Country', 'COUNTRY'),
            ('John', 25, 'USA', 'USA'),
            ('Alice', 30, 'Canada', 'CANADA'),
            ('Bob', 35, 'Australia', 'AUSTRALIA')
        ]
        self.assertEqual(processed_data, expected_processed_data)
        os.remove(output_file)

    def test_process_excel_data_4(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Age', 'COUNTRY'],
                ['John', 25, 'USA'],
                ['Alice', 30, 'CANADA'],
                ['Bob', 35, 'AUSTRALIA']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        N = 2
        success, output_file = processor.process_excel_data(N, self.test_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.isfile(output_file))
        processed_data = processor.read_excel(output_file)
        expected_processed_data = [
            ('Name', 'Age', 'COUNTRY', 'COUNTRY'),
            ('John', 25, 'USA', 'USA'),
            ('Alice', 30, 'CANADA', 'CANADA'),
            ('Bob', 35, 'AUSTRALIA', 'AUSTRALIA')
        ]
        self.assertEqual(processed_data, expected_processed_data)
        os.remove(output_file)

    def test_process_excel_data_5(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'AGE', 'COUNTRY'],
                ['John', 25, 'USA'],
                ['Alice', 30, 'CANADA'],
                ['Bob', 35, 'AUSTRALIA']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        N = 1
        success, output_file = processor.process_excel_data(N, self.test_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.isfile(output_file))
        processed_data = processor.read_excel(output_file)
        expected_processed_data = [
            ('Name', 'AGE', 'COUNTRY', 'AGE'),
            ('John', 25, 'USA', 25),
            ('Alice', 30, 'CANADA', 30),
            ('Bob', 35, 'AUSTRALIA', 35)
        ]
        self.assertEqual(processed_data, expected_processed_data)
        os.remove(output_file)

    def test_process_excel_data_6(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'AGE', 'COUNTRY'],
                ['John', 25, 'USA'],
                ['Alice', 30, 'CANADA'],
                ['Bob', 35, 'AUSTRALIA']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        res = processor.process_excel_data(100, self.test_file_name)
        self.assertEqual(res, 0)


class ExcelProcessorTest(unittest.TestCase):
    def test_ExcelProcessor(self):
        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Age', 'Country'],
                ['John', 25, 'USA'],
                ['Alice', 30, 'Canada'],
                ['Bob', 35, 'Australia']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        data = processor.read_excel(self.test_file_name)
        expected_data = [
            ('Name', 'Age', 'Country'),
            ('John', 25, 'USA'),
            ('Alice', 30, 'Canada'),
            ('Bob', 35, 'Australia')
        ]
        self.assertEqual(data, expected_data)

        processor = ExcelProcessor()
        new_data = [
            ('Name', 'Age', 'Country'),
            ('John', 25, 'USA'),
            ('Alice', 30, 'Canada'),
            ('Bob', 35, 'Australia'),
            ('Julia', 28, 'Germany')
        ]
        save_file_name = 'test_output.xlsx'
        success = processor.write_excel(new_data, save_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.exists(save_file_name))
        saved_data = processor.read_excel(save_file_name)
        self.assertEqual(saved_data, new_data)
        os.remove(save_file_name)

        self.test_file_name = 'test_data.xlsx'
        data = [['Name', 'Age', 'Country'],
                ['John', 25, 'USA'],
                ['Alice', 30, 'Canada'],
                ['Bob', 35, 'Australia']]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(self.test_file_name)
        workbook.close()

        processor = ExcelProcessor()
        N = 1
        success, output_file = processor.process_excel_data(N, self.test_file_name)
        self.assertTrue(success)
        self.assertTrue(os.path.isfile(output_file))
        processed_data = processor.read_excel(output_file)
        expected_processed_data = [
            ('Name', 'Age', 'Country', 'AGE'),
            ('John', 25, 'USA', 25),
            ('Alice', 30, 'Canada', 30),
            ('Bob', 35, 'Australia', 35)
        ]
        self.assertEqual(processed_data, expected_processed_data)
        os.remove(output_file)

